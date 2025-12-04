"""
SAD ALIGNMENT SCRIPT 
==============================================================
  ✔ normalizes Excel names → SAD-compatible names
  ✔ aligns ANY element that appears in Excel
  ✔ skips alignment ONLY when the element is not found in Excel

  -AA.
"""

import openpyxl
import re

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------
EXCEL_FILE       = "injector_lattice_20230303-2.xlsx"
EXCEL_SHEET      = "lattice"

SAD_INPUT_FILE   = "rfg-linac-bte.sad"
SAD_OUTPUT_FILE  = "lattice_with_alignment.sad"


# ------------------------------------------------------------
# NAME NORMALIZATION
# ------------------------------------------------------------
def normalize(name: str) -> str:
    """
    Convert both Excel and SAD names into a common form.
    Excel: QD_A1_G1 → QDA1G1
    """
    return name.upper().replace("_", "").replace("-", "").strip()


# ------------------------------------------------------------
# LOAD EXCEL ALIGNMENT TABLE
# ------------------------------------------------------------
print("\n Loading Excel alignment...")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[EXCEL_SHEET]

# Headers are on row 12
headers = [c.value for c in ws[12]]
col = {h: i for i, h in enumerate(headers)}

alignment = {}

for row in ws.iter_rows(min_row=13, values_only=True):
    raw_name = row[col["element_name"]]
    if not raw_name:
        continue

    key = normalize(raw_name)

    alignment[key] = {
        "DX":  (row[col["Δx [mm]"]] or 0) * 1e-3,
        "DY":  (row[col["Δz [mm]"]] or 0) * 1e-3,
        "ROT":  row[col["z' [rad]"]] or 0,
    }

print(f" Loaded {len(alignment)} alignment entries.\n")


# ------------------------------------------------------------
# SAD PARSING (INLINE ELEMENT DEFINITIONS)
# Example:
#    QUAD QDA1G1 =(L=0.0736 K1=-3.4)
#    DRIFT A=(L=0.1) B=(L=0.2)
# ------------------------------------------------------------

element_pattern = re.compile(r"([A-Za-z0-9_]+)\s*=\s*\(")

aligned_count = 0
missing_count = 0
missing_list = []

output_lines = []

print("Processing SAD lattice...\n")

with open(SAD_INPUT_FILE) as f:
    for line in f:
        modified_line = line
        matches = element_pattern.finditer(line)

        for m in matches:
            name = m.group(1)
            sad_key = normalize(name)

            # # Check Excel table
            # if sad_key not in alignment:
            #     missing_list.append(name)
            #     missing_count += 1
            #     continue

            if sad_key not in alignment:

                # Do NOT count missing if it starts with  (drifts, structures)
                if not (name.upper().startswith("L") or name.upper().startswith("G") or name.upper().startswith("W")  or name.upper().startswith("D") or name.upper().startswith("C") or name.upper().startswith("S") ):
                    missing_list.append(name)
                    missing_count += 1

                continue

            dx  = alignment[sad_key]["DX"]
            dy  = alignment[sad_key]["DY"]
            rot = alignment[sad_key]["ROT"]

            # Insert alignment parameters into SAD definition
            modified_line = modified_line.replace(
                m.group(0),
                f"{name} =(DX={dx:.6g} DY={dy:.6g} DROTATE={rot:.6g} "
            )

            aligned_count += 1

            print(f"  🔧 {name:12s}  DX={dx:.3g} DY={dy:.3g} ROT={rot:.3g}")

        output_lines.append(modified_line)


# ------------------------------------------------------------
# WRITE OUTPUT SAD FILE
# ------------------------------------------------------------
with open(SAD_OUTPUT_FILE, "w") as f:
    f.writelines(output_lines)


# ------------------------------------------------------------
# SUMMARY
# ------------------------------------------------------------
print("\n=====================================")
print(" ALIGNMENT UPDATE COMPLETE")
print("=====================================")
print(f" Elements aligned:      {aligned_count}")
print(f" Elements without data: {missing_count}")
print(f" Output written to:    {SAD_OUTPUT_FILE}")
print("=====================================\n")


# ------------------------------------------------------------
# SHOW FIRST 50 MISSING NAMES
# ------------------------------------------------------------
if missing_list:
    print("⚠ First missing element names (up to 50):")
    for name in missing_list[:50]:
        print("   ", name)
